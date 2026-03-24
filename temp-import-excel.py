"""
Import Excel FATURAMENTO data into contas_receber table on VPS.
Generates SQL INSERT statements and executes them remotely.
"""
import openpyxl
import paramiko
import os
import tempfile

# Excel config
EXCEL_PATH = os.path.join(os.path.dirname(__file__), 'Cópia de ALUFORCE - CR 2026 24-2 - Contas a Receber.xlsx')
SHEET_NAME = 'FATURAMENTO'
HEADER_ROW = 7
DATA_START_ROW = 8

# VPS config
VPS_HOST = '31.97.64.102'
VPS_USER = 'root'
VPS_PASS = 'Aluforce@2026#Vps'
DB_USER = 'aluforce'
DB_PASS = 'Aluforce2026VpsDB'
DB_NAME = 'aluforce_vendas'

def fmt_date(val):
    """Format date value to YYYY-MM-DD string"""
    if val is None:
        return None
    if hasattr(val, 'strftime'):
        return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    if not s:
        return None
    # Try to parse ISO format
    if len(s) >= 10:
        return s[:10]
    return None

def esc(val):
    """Escape string for MySQL"""
    if val is None:
        return 'NULL'
    s = str(val).strip()
    if not s:
        return 'NULL'
    # Escape single quotes and backslashes
    s = s.replace('\\', '\\\\').replace("'", "\\'")
    return f"'{s}'"

def fmt_num(val):
    """Format numeric value"""
    if val is None:
        return 'NULL'
    try:
        f = float(val)
        return f"{f:.2f}"
    except (ValueError, TypeError):
        return 'NULL'

def main():
    print("Reading Excel...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb[SHEET_NAME]
    
    # Column mapping (1-indexed): Excel col -> DB field
    # 1:EMPRESA, 2:EMISSÃO, 3:TIPO, 4:NFe, 5:P, 6:CLIENTE, 7:CNPJ, 
    # 8:VALOR P, 9:VCTO, 10:SITUAÇÃO, 11:PORTADOR, 12:DATA, 13:STATUS, 
    # 14:DIAS, 15:POSIÇÃO, 16:DATA2, 17:RECOMPRADO, 18:DATA3, 
    # 19:CARTORIO, 20:OBSERVAÇÃO, 21:ACEITA TROCA FACTORY, 22:COMISSÁRIA
    
    rows_data = []
    for row in range(DATA_START_ROW, ws.max_row + 1):
        empresa = ws.cell(row=row, column=1).value
        if not empresa:
            continue  # Skip empty rows
        
        emissao = fmt_date(ws.cell(row=row, column=2).value)
        tipo = ws.cell(row=row, column=3).value
        nfe = ws.cell(row=row, column=4).value
        parcela = ws.cell(row=row, column=5).value
        cliente = ws.cell(row=row, column=6).value
        cnpj = ws.cell(row=row, column=7).value
        valor = ws.cell(row=row, column=8).value
        vcto = fmt_date(ws.cell(row=row, column=9).value)
        situacao = ws.cell(row=row, column=10).value
        portador = ws.cell(row=row, column=11).value
        data_op = fmt_date(ws.cell(row=row, column=12).value)
        status_excel = ws.cell(row=row, column=13).value
        dias = ws.cell(row=row, column=14).value
        posicao = ws.cell(row=row, column=15).value
        data2 = fmt_date(ws.cell(row=row, column=16).value)
        recomprado = ws.cell(row=row, column=17).value
        data3 = fmt_date(ws.cell(row=row, column=18).value)
        cartorio = ws.cell(row=row, column=19).value
        observacao = ws.cell(row=row, column=20).value
        aceita_troca = ws.cell(row=row, column=21).value
        comissaria = ws.cell(row=row, column=22).value
        
        # Map Excel STATUS to DB status
        status_map = {
            'LIQUIDADO': 'recebido',
            'A VENCER': 'pendente',
            'VENCIDO': 'vencido',
        }
        db_status = 'pendente'
        if status_excel:
            db_status = status_map.get(str(status_excel).strip().upper(), 'pendente')
        
        # Calculate dias_vencido
        dias_int = None
        if dias is not None:
            try:
                dias_int = int(float(dias))
            except (ValueError, TypeError):
                dias_int = None
        
        rows_data.append({
            'empresa': empresa,
            'data_emissao': emissao,
            'tipo_documento_fiscal': tipo,
            'nota_fiscal': str(nfe) if nfe else None,
            'parcela_info': str(parcela) if parcela else None,
            'cliente_nome': cliente,
            'cnpj_cliente': cnpj,
            'valor': valor,
            'data_vencimento': vcto,
            'situacao': situacao,
            'portador': portador,
            'data_operacao': data_op,
            'status': db_status,
            'dias_vencido': dias_int,
            'posicao': posicao,
            'data_posicao': data2,
            'recomprado': recomprado,
            'data_recompra': data3,
            'cartorio': cartorio,
            'observacoes': observacao,
            'aceita_troca': aceita_troca,
            'comissaria': comissaria,
            'origem_importacao': 'excel',
        })
    
    print(f"Parsed {len(rows_data)} rows from Excel")
    
    # Generate SQL
    sql_lines = []
    sql_lines.append("SET NAMES utf8mb4;")
    sql_lines.append("-- Clear previous excel imports to avoid duplicates")
    sql_lines.append("DELETE FROM contas_receber WHERE origem_importacao = 'excel';")
    sql_lines.append("")
    
    # Build INSERT in batches of 50
    fields = [
        'empresa', 'data_emissao', 'tipo_documento_fiscal', 'nota_fiscal',
        'parcela_info', 'cliente_nome', 'cnpj_cliente', 'valor', 'data_vencimento',
        'situacao', 'portador', 'data_operacao', 'status', 'dias_vencido',
        'posicao', 'data_posicao', 'recomprado', 'data_recompra', 'cartorio',
        'observacoes', 'aceita_troca', 'comissaria', 'origem_importacao', 'descricao'
    ]
    
    batch_size = 50
    for i in range(0, len(rows_data), batch_size):
        batch = rows_data[i:i+batch_size]
        values_list = []
        for r in batch:
            vals = []
            for f in fields:
                if f == 'descricao':
                    # Set descricao = cliente_nome for display compatibility
                    vals.append(esc(r.get('cliente_nome')))
                elif f in ('data_emissao', 'data_vencimento', 'data_operacao', 'data_posicao', 'data_recompra'):
                    d = r.get(f)
                    vals.append(f"'{d}'" if d else 'NULL')
                elif f == 'valor':
                    vals.append(fmt_num(r.get(f)))
                elif f == 'dias_vencido':
                    d = r.get(f)
                    vals.append(str(d) if d is not None else 'NULL')
                else:
                    vals.append(esc(r.get(f)))
            values_list.append(f"({', '.join(vals)})")
        
        sql_lines.append(f"INSERT INTO contas_receber ({', '.join(fields)})")
        sql_lines.append(f"VALUES {','.join(values_list)};")
        sql_lines.append("")
    
    sql_content = '\n'.join(sql_lines)
    
    # Write SQL to temp file
    sql_file = os.path.join(os.path.dirname(__file__), 'temp-import-cr.sql')
    with open(sql_file, 'w', encoding='utf-8') as f:
        f.write(sql_content)
    print(f"Generated SQL file: {sql_file} ({len(sql_content)} bytes)")
    
    # Upload and execute on VPS
    print("Connecting to VPS...")
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    ssh.connect(VPS_HOST, username=VPS_USER, password=VPS_PASS)
    
    # Upload SQL file
    sftp = ssh.open_sftp()
    remote_sql = '/tmp/import-cr.sql'
    sftp.put(sql_file, remote_sql)
    sftp.close()
    print("SQL file uploaded")
    
    # Execute SQL
    cmd = f"mysql -u {DB_USER} -p{DB_PASS} {DB_NAME} < {remote_sql}"
    stdin, stdout, stderr = ssh.exec_command(cmd)
    out = stdout.read().decode()
    err = stderr.read().decode()
    if out:
        print("Output:", out[:500])
    if err and 'Warning' not in err:
        print("Error:", err[:500])
    elif err:
        print("(warnings only)")
    
    # Verify
    cmd2 = f"mysql -u {DB_USER} -p{DB_PASS} {DB_NAME} -e \"SELECT COUNT(*) as total FROM contas_receber; SELECT COUNT(*) as excel_rows FROM contas_receber WHERE origem_importacao='excel'; SELECT empresa, nota_fiscal, parcela_info, cliente_nome, situacao, portador, status, posicao FROM contas_receber WHERE origem_importacao='excel' LIMIT 3;\""
    stdin, stdout, stderr = ssh.exec_command(cmd2)
    print("\nVerification:")
    print(stdout.read().decode())
    
    # Cleanup
    ssh.exec_command(f"rm {remote_sql}")
    ssh.close()
    print("Import complete!")

if __name__ == '__main__':
    main()
