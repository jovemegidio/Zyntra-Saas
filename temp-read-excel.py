import openpyxl
wb = openpyxl.load_workbook(r'Cópia de ALUFORCE - CR 2026 24-2 - Contas a Receber.xlsx', data_only=True, read_only=True)

print('=== FUNDOS ===')
ws = wb['FUNDOS']
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=25, values_only=True)):
    vals = [str(v) if v is not None else '' for v in row]
    print('R%d: %s' % (i+1, ' | '.join(vals)))

print()
print('=== DIARIO (first 15 rows) ===')
ws = wb['DIARIO']
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True)):
    vals = [str(v) if v is not None else '' for v in row]
    print('R%d: %s' % (i+1, ' | '.join(vals)))

print()
print('=== PM (first 10) ===')
ws = wb['PM']
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True)):
    vals = [str(v) if v is not None else '' for v in row]
    print('R%d: %s' % (i+1, ' | '.join(vals)))

print()
print('=== TD ===')
ws = wb['TD']
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True)):
    vals = [str(v) if v is not None else '' for v in row]
    print('R%d: %s' % (i+1, ' | '.join(vals)))
