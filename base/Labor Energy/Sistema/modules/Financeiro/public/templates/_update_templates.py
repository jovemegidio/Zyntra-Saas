# -*- coding: utf-8 -*-
"""
Atualiza os templates XLSX do modulo Financeiro:
- Logo ALUFORCE compacta (estilo Omie)
- Titulo ao lado da logo
- Cores conforme modulo (Financeiro=verde, Vendas=azul, Compras=roxo)
- Layout baseado no padrao SGE/Omie
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XlImage
import os
import shutil

# Config
TEMPLATES_DIR = os.path.dirname(os.path.abspath(__file__))
LOGO_PATH = os.path.join(TEMPLATES_DIR, '..', '..', '..', '..', 'public', 'images', 'Logo Monocromatico - Azul - Aluforce.png')
LOGO_PATH = os.path.normpath(LOGO_PATH)

# Cores por modulo
COLORS = {
    'Financeiro': {'primary': '16a34a', 'light': 'dcfce7', 'bg': 'f0fdf4'},
    'Vendas':     {'primary': '0a4f7e', 'light': 'bfdbfe', 'bg': 'eff6ff'},
    'Compras':    {'primary': '7c3aed', 'light': 'ddd6fe', 'bg': 'f5f3ff'},
}

HEADER_ROWS = 4  # logo+title, subtitle, instruction, blank separator

TEMPLATES = {
    'template_contas_pagar.xlsx': {
        'titulo': 'Planilha de Importacao de Contas a Pagar',
        'subtitulo': 'Modulo Financeiro',
        'instrucao': 'Preencha os dados a partir da linha 6. Campos com * sao obrigatorios. Nao altere o cabecalho.',
        'modulo': 'Financeiro',
    },
    'template_contas_receber.xlsx': {
        'titulo': 'Planilha de Importacao de Contas a Receber',
        'subtitulo': 'Modulo Financeiro',
        'instrucao': 'Preencha os dados a partir da linha 6. Campos com * sao obrigatorios. Nao altere o cabecalho.',
        'modulo': 'Financeiro',
    },
    'template_bancos.xlsx': {
        'titulo': 'Planilha de Importacao de Contas Bancarias',
        'subtitulo': 'Modulo Financeiro',
        'instrucao': 'Preencha os dados a partir da linha 6. Campos com * sao obrigatorios. Nao altere o cabecalho.',
        'modulo': 'Financeiro',
    },
    'template_movimentacoes.xlsx': {
        'titulo': 'Planilha de Importacao de Movimentacoes',
        'subtitulo': 'Modulo Financeiro',
        'instrucao': 'Preencha os dados a partir da linha 6. Campos com * sao obrigatorios. Nao altere o cabecalho.',
        'modulo': 'Financeiro',
    },
    'SGE_Fluxo_Caixa.xlsx': {
        'titulo': 'Planilha de Importacao de Fluxo de Caixa',
        'subtitulo': 'Modulo Financeiro',
        'instrucao': 'Preencha os dados a partir da linha 6. Campos com * sao obrigatorios. Nao altere o cabecalho.',
        'modulo': 'Financeiro',
    },
}

WHITE = 'FFFFFF'


def get_styles(modulo):
    """Retorna estilos baseados no modulo"""
    colors = COLORS.get(modulo, COLORS['Financeiro'])
    primary = colors['primary']

    return {
        'title_font': Font(name='Calibri', size=14, bold=True, color=WHITE),
        'subtitle_font': Font(name='Calibri', size=10, color=WHITE),
        'instruction_font': Font(name='Calibri', size=9, italic=True, color='666666'),
        'header_font': Font(name='Calibri', size=10, bold=True, color=WHITE),
        'header_fill': PatternFill(start_color=primary, end_color=primary, fill_type='solid'),
        'title_fill': PatternFill(start_color=primary, end_color=primary, fill_type='solid'),
        'example_font': Font(name='Calibri', size=10, color='999999', italic=True),
        'alt_fill': PatternFill(start_color=colors['bg'], end_color=colors['bg'], fill_type='solid'),
        'thin_border': Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC'),
        ),
    }


def update_template(filename, config):
    filepath = os.path.join(TEMPLATES_DIR, filename)
    if not os.path.exists(filepath):
        print(f'  SKIP: {filename} not found')
        return False

    backup = filepath + '.bak'
    shutil.copy2(filepath, backup)

    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    max_col = ws.max_column
    max_row = ws.max_row

    styles = get_styles(config['modulo'])

    # Step 1: Insert header rows
    ws.insert_rows(1, HEADER_ROWS)

    # Step 2: Row 1 - Logo compacta + Title (fundo colorido do modulo)
    ws.row_dimensions[1].height = 40

    # Logo compacta (36px altura - estilo Omie)
    if os.path.exists(LOGO_PATH):
        img = XlImage(LOGO_PATH)
        ratio = img.width / img.height if img.height > 0 else 1
        img.height = 36
        img.width = int(36 * ratio)
        ws.add_image(img, 'A1')

    # Title ao lado da logo (coluna C em diante)
    title_col = min(3, max_col)
    title_cell = ws.cell(row=1, column=title_col, value=config['titulo'])
    title_cell.font = styles['title_font']
    title_cell.alignment = Alignment(vertical='center')

    # Fundo colorido em toda row 1
    for col in range(1, max_col + 1):
        ws.cell(row=1, column=col).fill = styles['title_fill']

    # Merge titulo
    if max_col > title_col:
        ws.merge_cells(start_row=1, start_column=title_col, end_row=1, end_column=max_col)

    # Step 3: Row 2 - Subtitle (fundo colorido)
    ws.row_dimensions[2].height = 20
    sub_cell = ws.cell(row=2, column=title_col, value=config['subtitulo'])
    sub_cell.font = styles['subtitle_font']
    sub_cell.alignment = Alignment(vertical='center')
    for col in range(1, max_col + 1):
        ws.cell(row=2, column=col).fill = styles['title_fill']
    if max_col > title_col:
        ws.merge_cells(start_row=2, start_column=title_col, end_row=2, end_column=max_col)

    # Step 4: Row 3 - Instructions (fundo branco)
    ws.row_dimensions[3].height = 18
    inst_cell = ws.cell(row=3, column=1, value=config['instrucao'])
    inst_cell.font = styles['instruction_font']
    inst_cell.alignment = Alignment(vertical='center')
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=max_col)

    # Step 5: Row 4 - Blank separator
    ws.row_dimensions[4].height = 4

    # Step 6: Style column headers (row 5)
    header_row = HEADER_ROWS + 1
    ws.row_dimensions[header_row].height = 28
    for col in range(1, max_col + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = styles['header_font']
        cell.fill = styles['header_fill']
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = styles['thin_border']

    # Step 7: Style data rows
    for row in range(header_row + 1, header_row + 1 + max_row):
        is_alt = (row - header_row) % 2 == 0
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                cell.font = styles['example_font']
                cell.border = styles['thin_border']
                if is_alt:
                    cell.fill = styles['alt_fill']
                cell.alignment = Alignment(vertical='center')

    # Step 8: Auto-fit column widths
    for col in range(1, max_col + 1):
        max_len = 0
        col_letter = get_column_letter(col)
        for row in range(header_row, header_row + 1 + max_row):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        width = min(max(max_len + 2, 12), 35)
        ws.column_dimensions[col_letter].width = width

    # Step 9: Freeze panes
    ws.freeze_panes = f'A{header_row + 1}'

    wb.save(filepath)
    wb.close()
    os.remove(backup)
    return True


def main():
    print(f'Logo path: {LOGO_PATH}')
    print(f'Logo exists: {os.path.exists(LOGO_PATH)}')
    print(f'Templates dir: {TEMPLATES_DIR}')
    print()

    ok = 0
    for filename, config in TEMPLATES.items():
        print(f'Processing {filename}...')
        if update_template(filename, config):
            print(f'  OK')
            ok += 1
        else:
            print(f'  FAILED')

    print(f'\nDone: {ok}/{len(TEMPLATES)} templates updated')


if __name__ == '__main__':
    main()
